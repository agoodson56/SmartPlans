#!/usr/bin/env python3
"""Inspect specific sheets of a bid to find the total cell."""
import sys
import openpyxl
import xlrd
from pathlib import Path

def inspect(path, sheet_name=None, max_rows=80):
    if path.suffix.lower() == ".xls":
        wb = xlrd.open_workbook(str(path))
        sheets = wb.sheets()
        if sheet_name:
            sheets = [s for s in sheets if s.name == sheet_name]
        for sh in sheets:
            print(f"\n=== Sheet: {sh.name} ({sh.nrows} rows × {sh.ncols} cols) ===")
            for r in range(min(sh.nrows, max_rows)):
                cells = [sh.cell_value(r, c) for c in range(sh.ncols)]
                # Skip blank rows
                if all(not c or (isinstance(c, str) and not c.strip()) for c in cells):
                    continue
                # Trim trailing blanks
                while cells and (cells[-1] == "" or cells[-1] is None):
                    cells.pop()
                if not cells:
                    continue
                line = " | ".join(repr(c)[:40] for c in cells)
                print(f"  R{r:>3}: {line}")
    else:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheets = wb.worksheets
        if sheet_name:
            sheets = [s for s in sheets if s.title == sheet_name]
        for ws in sheets:
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            print(f"\n=== Sheet: {ws.title} (~{row_count} rows × {col_count} cols) ===")
            r = 0
            for row in ws.iter_rows(values_only=True):
                r += 1
                if r > max_rows:
                    break
                if not row:
                    continue
                cells = list(row)
                while cells and (cells[-1] == "" or cells[-1] is None):
                    cells.pop()
                if not cells:
                    continue
                # Skip blank rows
                if all(not c or (isinstance(c, str) and not c.strip()) for c in cells):
                    continue
                line = " | ".join(repr(c)[:40] for c in cells)
                print(f"  R{r:>3}: {line}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: inspect-bid.py <file> [sheet]")
        sys.exit(1)
    p = Path(sys.argv[1])
    sn = sys.argv[2] if len(sys.argv) > 2 else None
    inspect(p, sn)
